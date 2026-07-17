"""
01_extract_data.py
Extract monthly NSE equity-derivatives and cash-market series from SEBI
Monthly Bulletin annexure tables (Excel), and hand-code the cohort tables
from SEBI's July 2025 study.

Sources (all public, sebi.gov.in):
  - SEBI Bulletin April 2023 (Excel): FY2022-23 monthly EDS (notional only)
  - SEBI Bulletin April 2024 (Excel): FY2023-24 monthly EDS (premium + notional)
  - SEBI Bulletin March 2025 (Excel): FY2024-25 monthly EDS Apr-24..Feb-25
  - SEBI Bulletin April 2025 (PDF)  : Mar-25 row (hand-coded below)
  - SEBI Bulletin April 2026 (Excel): FY2025-26 monthly EDS
  - SEBI (Jul 2025): "Comparative study of growth in EDS vis-a-vis Cash Market"
"""
import openpyxl, warnings, datetime as dt
import pandas as pd
import numpy as np

warnings.filterwarnings("ignore")
DATA = "/sessions/serene-great-cannon/mnt/outputs/data/"

# ---------------------------------------------------------------- helpers
def month_key(v):
    """Normalize month cell (datetime or 'Apr-22' string) -> pd.Period."""
    if isinstance(v, dt.datetime):
        return pd.Period(v, freq="M")
    s = str(v).strip()
    for fmt in ("%b-%y", "%b-%Y"):
        try:
            return pd.Period(dt.datetime.strptime(s, fmt), freq="M")
        except ValueError:
            pass
    return None

def rows(f, sheet):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    return list(wb[sheet].iter_rows(values_only=True))

def num(x):
    if x is None: return np.nan
    try: return float(str(x).replace(",", ""))
    except ValueError: return np.nan

FILES = ["sebi_bull_apr2023.xlsx", "sebi_bull_apr2024.xlsx",
         "sebi_bull_mar2025.xlsx", "sebi_bull_apr2026.xlsx"]

def find_sheet(f, title_frag):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    for s in wb.sheetnames:
        for row in wb[s].iter_rows(max_row=1, values_only=True):
            if row and row[0] and title_frag in str(row[0]):
                return s
    return None

def extract_eds(exchange):
    """exchange in {'NSE','BSE'} -> monthly frame from the 4 bulletin files."""
    rec = []
    for f in FILES:
        sheet = find_sheet(DATA + f, f"Equity Derivatives Segment at {exchange}")
        assert sheet, (f, exchange)
        old_format = "apr2023" in f
        for r in rows(DATA + f, sheet):
            m = month_key(r[0])
            if m is None: continue
            if old_format:  # FY23: call/put notional only
                rec.append(dict(month=m, days=num(r[1]),
                                idx_fut=num(r[3]), stk_fut=num(r[5]),
                                idxopt_prem=np.nan,
                                idxopt_notional=num(r[7]) + num(r[9]),
                                stkopt_prem=np.nan,
                                stkopt_notional=num(r[11]) + num(r[13])))
            else:           # FY24+: premium + notional
                rec.append(dict(month=m, days=num(r[1]),
                                idx_fut=num(r[3]), stk_fut=num(r[5]),
                                idxopt_prem=num(r[7]) + num(r[10]),
                                idxopt_notional=num(r[8]) + num(r[11]),
                                stkopt_prem=num(r[13]) + num(r[16]),
                                stkopt_notional=num(r[14]) + num(r[17])))
    df = pd.DataFrame(rec).dropna(subset=["days"])
    return df[~df.month.isna()].drop_duplicates("month").set_index("month").sort_index()

nse = extract_eds("NSE")
bse = extract_eds("BSE")

# ---- Mar-2025 rows (SEBI Bulletin April 2025 PDF, Tables 32-33) -------------
fy25 = [pd.Period(f"2024-{m:02d}", "M") for m in range(4, 13)] + \
       [pd.Period(f"2025-{m:02d}", "M") for m in (1, 2)]

mar25_nse = dict(days=19.0, idx_fut=586218.33, stk_fut=2390586.99,
                 idxopt_prem=453174.94 + 390937.10,
                 idxopt_notional=214715084.65 + 198657722.40)
mar25_nse["stkopt_prem"] = (1305390.77 + 669801.70) - nse.loc[fy25, "stkopt_prem"].sum()
mar25_nse["stkopt_notional"] = (88222586.46 + 42894740.52) - nse.loc[fy25, "stkopt_notional"].sum()
nse.loc[pd.Period("2025-03", "M")] = mar25_nse

mar25_bse = dict(days=19.0, idx_fut=4259.57, stk_fut=76.14,
                 idxopt_prem=122718.12 + 113483.08,
                 idxopt_notional=125443391.50 + 118936113.02)
mar25_bse["stkopt_prem"] = (10.48 + 8.90) - bse.loc[fy25, "stkopt_prem"].sum()
mar25_bse["stkopt_notional"] = (1970.40 + 630.93) - bse.loc[fy25, "stkopt_notional"].sum()
bse.loc[pd.Period("2025-03", "M")] = mar25_bse

nse, bse = nse.sort_index(), bse.sort_index()

# ---- Market-wide (NSE+BSE) ---------------------------------------------------
cols = ["idx_fut", "stk_fut", "idxopt_prem", "idxopt_notional",
        "stkopt_prem", "stkopt_notional"]
mkt = nse[cols].add(bse[cols].reindex(nse.index).fillna(0.0))
# FY23 premiums are NaN in both -> keep NaN (fillna(0) on bse only affects values where nse notnull)
mkt.loc[nse["idxopt_prem"].isna(), ["idxopt_prem", "stkopt_prem"]] = np.nan
mkt["days"] = nse["days"]  # NSE trading days as market convention
for c in cols:
    mkt[c + "_adt"] = mkt[c] / mkt["days"]
    nse[c + "_adt"] = nse[c] / nse["days"]

# ------------------------------------------------- Cash segments
def extract_cash(exchange, mar25_row):
    rec = []
    for f in FILES:
        sheet = find_sheet(DATA + f, f"Trends in Cash Segment of {exchange}")
        assert sheet, (f, exchange)
        for r in rows(DATA + f, sheet):
            m = month_key(r[0])
            if m is None: continue
            rec.append(dict(month=m, cash_days=num(r[4]),
                            cash_turnover=num(r[7]), cash_adt=num(r[8])))
    df = pd.DataFrame(rec).dropna(subset=["cash_turnover"])
    df = df.drop_duplicates("month").set_index("month").sort_index()
    df.loc[pd.Period("2025-03", "M")] = mar25_row
    return df.sort_index()

cash_nse = extract_cash("NSE", dict(cash_days=19, cash_turnover=1875160.0, cash_adt=98693.0))
cash_bse = extract_cash("BSE", dict(cash_days=19, cash_turnover=107037.0, cash_adt=5634.0))
cash = cash_nse.copy()
cash["cash_turnover"] += cash_bse["cash_turnover"].reindex(cash.index).fillna(0)
cash["cash_adt"] = cash["cash_turnover"] / cash["cash_days"]

panel = mkt.join(cash, how="left")
panel.index.name = "month"
panel.to_csv(DATA + "monthly_market_panel.csv")
nse.join(cash_nse, how="left").to_csv(DATA + "monthly_nse_panel.csv")
bse.to_csv(DATA + "monthly_bse_panel.csv")

print(panel[["days", "idxopt_prem_adt", "idxopt_notional_adt", "stkopt_prem_adt",
             "idx_fut_adt", "stk_fut_adt", "cash_adt"]].tail(18).round(0).to_string())
print("\nrows:", len(panel), "| range:", panel.index.min(), "->", panel.index.max())

# --- sanity check vs SEBI Table 6 (Dec-24..May-25 market-wide index options)
w = [pd.Period(x, "M") for x in
     ["2024-12", "2025-01", "2025-02", "2025-03", "2025-04", "2025-05"]]
chk_prem = panel.loc[w, "idxopt_prem"].sum() / panel.loc[w, "days"].sum()
chk_not = panel.loc[w, "idxopt_notional"].sum() / panel.loc[w, "days"].sum()
print(f"\nDec24-May25 idx-opt ADT premium {chk_prem:,.0f} (SEBI: 61,533)")
print(f"Dec24-May25 idx-opt ADT notional {chk_not:,.0f} (SEBI: 3,18,50,658)")

# ---------------------------------------------- SEBI Jul-2025 study tables
# Table 8: unique traders by turnover bucket, three Dec-May windows
t8 = pd.DataFrame({
    "bucket": ["<10k", "10k-1L", "1L-10L", "10L-1Cr", "1Cr-10Cr", ">10Cr"],
    "traders_2223": [659746, 935569, 1494229, 1505975, 710580, 178891],
    "traders_2324": [1482603, 1601235, 2171763, 1984051, 935833, 249799],
    "traders_2425": [1036412, 1247137, 1661303, 1708985, 898435, 222202],
})
t8.to_csv(DATA + "sebi_t8_cohorts.csv", index=False)

# Table 12: quarterly FY25 individual traders P&L
t12 = pd.DataFrame({
    "quarter": ["FY25Q1", "FY25Q2", "FY25Q3", "FY25Q4"],
    "net_profit_cr": [-21255, -25942, -33661, -24745],
    "traders_lakh": [61.4, 59.2, 53.5, 42.7],
    "loss_makers_pct": [84.5, 86.3, 88.5, 86.4],
    "avg_pl": [-34606, -43847, -62975, -57920],
})
t12.to_csv(DATA + "sebi_t12_quarterly.csv", index=False)

# Table 11: annual P&L FY22-FY25
t11 = pd.DataFrame({
    "fy": ["FY22", "FY23", "FY24", "FY25"],
    "net_profit_cr": [-40824, -65747, -74812, -105603],
    "traders_lakh": [42.7, 58.4, 86.3, 96.0],
    "loss_makers_pct": [90.2, 91.7, 91.1, 91.0],
    "avg_pl": [-95517, -112677, -86728, -110069],
})
t11.to_csv(DATA + "sebi_t11_annual.csv", index=False)
print("\nSEBI study tables written.")
